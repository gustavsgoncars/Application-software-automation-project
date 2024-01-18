import selenium
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
import time
from datetime import datetime
from openpyxl import Workbook, load_workbook
from selenium.common.exceptions import NoSuchElementException
import unidecode
from unidecode import unidecode
import re

todays_date = datetime.now()
today = todays_date.day # iegūst programmas palaišanas dienas datumu 

x = input("Studiju kursa nosaukums: ")
x = unidecode(x)
x = x.lower()
while True:
    try:
        y = int(input("Kurss: "))
        y += 1
        break  # ja viss ir kārtībā, tad tiek cikls pārtraukts
    except ValueError:
        print("Nepareiza vērtība. Jums ir jāievada vēlamā kursa nr. bez punkta vai atstarpēm. Lūdzu, nēģiniet vēlreiz.")
while True:
    try:
        z = int(input("Grupa: "))
        break  # ja viss ir kārtībā, tad tiek cikls pārtraukts
    except ValueError:
        print("Nepareiza vērtība. Jums ir jāievada vēlamās grupas nr. bez punkta vai atstarpēm. Lūdzu, mēģiniet vēlreiz.")

while True:
    same_loc = input("Vai jūs šomēness VIENMĒR izbrauksiet no tās pašas adreses? (atbilde - y/n) ").lower()
    if same_loc == 'y' or same_loc == 'n':
        break
    else:
        print("Lūdzu, ievadiet derīgu atbildi ('y' vai 'n').")

if same_loc=="y":
    while True:
        Adrese = input("Kāda ir adrese no kuras jūs izbrauksiet? ")
        if Adrese: # nevar pārbaudīt vai ir ievadīta eksistējoša adrese, bet var pārbaudīt vai lietotājs vispār kaut ko ir ievadījis
            break
        else:
            print("Lūdzu ievadiet adresi.")
elif same_loc=="n":
    Adrese="n"

while True:
    try:
        e = input("Lūdzu ievadiet, kādā valodā izmantosiet pārlūku.('lv' vai 'eng') ")
        break  # ja viss ir kārtībā, tad tiek cikls pārtraukts
    except ValueError:
        print("Nepareiza vērtība. Jums ir jāievada 'lv' pārlūkam latviešu valodā vai 'eng' pārlūkam angļu valodā.")

service = Service()
option = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=option)

url = "https://nodarbibas.rtu.lv/"
driver.get(url)
time.sleep(1)

find = driver.find_element(By.XPATH, "/html/body/div[2]/div/div[1]/div/div/div/div[2]/div/button") # nospiež uz studiju kursa nosaukuma izvēli
find.click()

find.send_keys(x) # ievada kursa nosaukumu
find.send_keys(Keys.ENTER) # nospiež ENTER tādā veidā izēloties vistuvāk lietotāja input atbilstošo kursu

find = driver.find_element(By.XPATH, "/html/body/div[2]/div/div[1]/div/div/div/div[3]/div[1]/div[1]/select") # nospiež kursa nr. izvēli
find.click()
time.sleep(1)

find = driver.find_element(By.XPATH, "/html/body/div[2]/div/div[1]/div/div/div/div[3]/div[1]/div[1]/select/option[" + str(y) + "]") 
# tiek nospiests kurss atbilstoši user input
find.click()
time.sleep(1)

find = driver.find_element(By.XPATH, "/html/body/div[2]/div/div[1]/div/div/div/div[3]/div[1]/div[2]/select") # nospiež grupas nr. izvēli
find.click()
time.sleep(1)

select = Select(find)
select.select_by_visible_text(str(z)) # tiek aplūkota redzamā informācija starp grupu opcijām un no redzamām opcijām izvēlēta tā, kas atbilst user input
time.sleep(1)

sections = driver.find_elements(By.CLASS_NAME, 'fc-daygrid-day') # sections atradīs katras kalendārās dienas elementus
datetime = []

prev_d = 0
for s in sections:
    try:
        date_elements = s.find_elements(By.CLASS_NAME, 'fc-daygrid-day-number') # dienas datumi
        time_elements = s.find_elements(By.CLASS_NAME, 'fc-event-time') # lekciju laiki
        location_elements = s.find_elements(By.CLASS_NAME, 'fc-event-title') # lekciju norises vieta
        
        date=[]
        for d in date_elements:
            date.append(d.text)
            clean_time = []
            locations = []

            for l, t in zip(location_elements, time_elements):
                location_text = l.text
                location_text = location_text[-30:]
                match = re.search(r'\(([^-]*)-', location_text) # mēs zinam, ka adresi ierobežo ( un - simboli, tad mēs paņemam tekstu, kas ir starp šiem simboliem 
                if match:
                    cleaned_loc = match.group(1)
                    if cleaned_loc != "Att. ": # nav vajdzīgs iekļaut sarakstā lekcijas, kuras notiek attālināti
                        cleaned_loc = cleaned_loc.split(' ') # Atdalam ielas saīsinājumu no ielas nummura tālākai apstrādei
                        locations.append(cleaned_loc)
                        t = t.text
                        split_time = t.replace(' ', '').replace(':', '').split('-') # satīra laika datus
                        clean_time.extend(split_time)

            if clean_time != [] and locations != [] and date != []: 
                start_time = min(clean_time)
                end_time = max(clean_time)
                datetime.append({"dates": date, "locations": locations, "start": start_time, "end": end_time}) # tiek izeidots saraksts, kas sastav no 4triem dictionaries
  
    except NoSuchElementException:
        continue

driver.quit() # pabeigts darbs ar nodarbibas.rtu.lv

wb1=load_workbook('nos.xlsx')
wb1.create_sheet('Sheet2')
ws1_1=wb1['Sheet1']
ws1_2=wb1['Sheet2']
max_row1=len(datetime)
wb2=load_workbook('ielas.xlsx') # šajā failā ir visi pilnie ielu nosaukumi
ws2=wb2.active
max_row2=ws2.max_row

for b in range(0, len(datetime)):
    loc_start = datetime[b]["locations"][0][0] # pirmās lekcijas adreses saīsinājums
    loc_end = datetime[b]["locations"][-1][0] # pēdējās lekcijas adreses saīsinājums
    for line in range(1, max_row2+1):
        full_street = str(ws2['B'+str(line)].value) # ielu pilnais nosaukums
        ab_street = str(ws2['A'+str(line)].value) # ielu saīsinājumi excel failā
        if ab_street == loc_end:
            datetime[b]["locations"][-1][0] = full_street + " " + datetime[b]["locations"][-1][1] # pārveidojam uz pilno nosaukumu
        if ab_street == loc_start:
            datetime[b]["locations"][0][0] = full_street + " " + datetime[b]["locations"][0][1] # pārveidojam uz pilno nosaukumu

wb2.close() # aizver excel failu ar garajiem ielu nosaukumiem 

a=0
for row in range(2,max_row1+2): # savadam visus iegūtos datus pirmajā darba lapā
    day=datetime[a]["dates"][0]
    start=datetime[a]["start"]
    end=datetime[a]["end"]
    loc_start=datetime[a]["locations"][0][0]
    loc_end = datetime[a]["locations"][-1][0]
    a+=1
    ws1_1['A'+str(row)].value=day #dienas A kolonna
    ws1_1['B'+str(row)].value=start #lekcijas sākuma laiks B kolonna
    ws1_1['C'+str(row)].value=loc_start#lekcijas sākuma vieta C kolonna
    ws1_1['D'+str(row)].value=end ##lekcijas beigu laiks D kolonna
    ws1_1['E'+str(row)].value=loc_end#lekcijas beigu vieta E kolonna

service = Service()
option = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=option)    

url = "https://www.google.com/maps"
driver.get(url)
time.sleep(1)

# reject cookies button
find = driver.find_element(By.XPATH, "/html/body/c-wiz/div/div/div/div[2]/div[1]/div[3]/div[1]/div[1]/form[1]/div/div/button")
find.click()
time.sleep(1)

if str(e) == "lv":
    cycle=6
if str(e) == "eng":
    cycle=5
start_row=2
row_count_keeper=2
for r in range(2,max_row1+2):
    arrive_by_time=ws1_1['B'+str(r)].value
    arrive_at_loc=ws1_1['C'+str(r)].value
    cycle_day=ws1_1['A'+str(r)].value

    if int(cycle_day) < today:
        continue

    find = driver.find_element(By.XPATH, "/html/body/div[1]/div[3]/div[8]/div[3]/div[1]/div[1]/div/div[2]/form/input") # input lauks
    find.clear()
    find.send_keys(arrive_at_loc) # ievada adresi uz kuru tiks aprēķināts maršruts
    find.send_keys(Keys.ENTER)
    time.sleep(5)

    find = driver.find_element(By.XPATH, "/html/body/div[1]/div[3]/div[8]/div[9]/div/div/div[1]/div[2]/div/div[1]/div/div/div[4]/div[1]/button") # poga "marsruts"/"directions"
    find.click()
    time.sleep(5)

    # ievadam adresi no kuras izbrauks
    if Adrese=="n": # ja šis izpildās user is asked katrā loop no kurienes vins taja dienā izbrauks
        print(cycle_day + ". izbraukšanas adrese ir:")
        from_loc=input()
        find = driver.find_element(By.XPATH, "/html/body/div[1]/div[3]/div[8]/div[3]/div[1]/div[2]/div/div[3]/div[1]/div[1]/div[2]/div[1]/div/input")
        find.click()
        find.clear()
        find.send_keys(from_loc)
        time.sleep(5)
    else: # citādāk uzskatam, ka ir ievadīta konstanta adrese, ko mēs izmantosim visās iterācijās
        from_loc=Adrese
        find = driver.find_element(By.XPATH, "/html/body/div[1]/div[3]/div[8]/div[3]/div[1]/div[2]/div/div[3]/div[1]/div[1]/div[2]/div[1]/div/input")
        find.click()
        find.clear()
        find.send_keys(from_loc)
        time.sleep(3)
    
    find = driver.find_element(By.XPATH, "/html/body/div[1]/div[3]/div[8]/div[3]/div[1]/div[2]/div/div[2]/div/div/div/div[1]/button") # nospiez "best travel" braukšanas pogu, sis ir vajadzīgs, lai ciks sākot ar otro iterāciju nesāktu bufferot
    find.click()
    time.sleep(3)

    find = driver.find_element(By.XPATH, "/html/body/div[1]/div[3]/div[8]/div[3]/div[1]/div[2]/div/div[2]/div/div/div/div[3]/button") # izbraukt ar autobusu poga
    find.click()
    time.sleep(3)
    
    find = driver.find_element(By.XPATH, "/html/body/div[1]/div[3]/div[8]/div[9]/div/div/div[1]/div[2]/div/div[1]/div/div/div[2]/span/div") # "Leave now" poga
    find.click()
    time.sleep(4)

    find = driver.find_element(By.XPATH, "/html/body/div[" + str(cycle) + "]/div[3]/div") # "Arrive by" poga
    find.click()
    time.sleep(2)
    
    find = driver.find_element(By.XPATH, "/html/body/div[1]/div[3]/div[8]/div[9]/div/div/div[1]/div[2]/div/div[1]/div/div/div[2]/div/span[1]/input") # time ievade
    find.clear()
    find.send_keys(arrive_by_time)
    time.sleep(3)

    check_cycle_day=int(cycle_day)
    while check_cycle_day > today: # ar pogu "next" aizejam līdz vajadzīgajai dienai 
        # (izvēlēties datumu caur GoogleMaps kalendāru nebija iespējams, līdzko gāja otrā iterācija tā datumi palika "not interactable element" un to apiet nekādīgi nesanāca)
        find = driver.find_element(By.XPATH, "/html/body/div[1]/div[3]/div[8]/div[9]/div/div/div[1]/div[2]/div/div[1]/div/div/div[2]/div/span[2]/span[3]/button[2]") # uz prieksu poga
        find.click()
        time.sleep(2)
        check_cycle_day -= 1
    
    find = driver.find_element(By.CSS_SELECTOR, "[data-trip-index='0']") # izvēlamies pirmo piedāvāto maršrutu
    find.click()
    time.sleep(5)

    h2_loc=driver.find_elements(By.TAG_NAME, 'h2') # nosaukumi pieturam, izkapšanam utt
    h2_values = [element.text for element in h2_loc if element.text.strip() != ''] # izveidojas list, tiek izslēgti tukšie ieraksti

    class_times = driver.find_elements(By.CLASS_NAME, 'nTib6e') # cikos tu busi noteiktā punktā
    time_values = [element.text.replace("\u202f", " ") for element in class_times if element.text.strip() != '']

    class_desc = driver.find_elements(By.CSS_SELECTOR, "[class*='JxBYrc pk9Qwb']") # walk vai sabiedriskā nosaukums
    desc_values = [element.text for element in class_desc if element.text.strip() != '']
    
    tr_type = driver.find_elements(By.CSS_SELECTOR, "[class*='V0uhjd']") # iegūstam informāciju, kas būs vajadzīga, lai noteiktu transporta veidu
    el_style = []
    for element in tr_type:
        style_attribute = element.get_attribute("style")
        el_style.append(style_attribute if style_attribute else ' ') # ja nav style attribute tad ieliekam sarakstā tukšu elementu, lai pēc tam būtu vieglāk ievadīt informāciju excel

    for i in range(len(el_style)): # noteikta krāsa atbilst noteiktam transporta veidam, tātad mēs pārveidojam krāsu uz transporta veidu sarakstā
        if el_style[i] == "border-color: rgb(242, 179, 39);":
            el_style[i] = "Autobuss"
        if el_style[i] == "border-color: rgb(165, 167, 168);":
            el_style[i] = "Mikroautobuss"
        if el_style[i] == "border-color: rgb(0, 157, 224);":
            el_style[i] = "Trolejbuss"
        if el_style[i] == "border-color: rgb(242, 0, 11);":
            el_style[i] = "Tramvajs"
        if el_style[i] == "border-color: rgb(141, 3, 78);":
            el_style[i] = "Mikroautobuss"
        if el_style[i] == "border-color: rgb(139, 197, 64);" or el_style[i] == "border-color: rgb(234, 100, 73);" or el_style[i] == "border-color: rgb(234, 100, 73);":
            el_style[i] = "Vilciens"


    class_numbers = driver.find_element(By.CLASS_NAME, 'CMnFh') # sabiedrisko transportu numurs 
    class_numbers = class_numbers.text
    numbers_values = class_numbers.split(" ")
    numbers_values = [element for element in numbers_values if element.strip() != '\ue315'] # paturam tikai numuru

    walk_transp = []
    if str(e) == "lv":
        for i, element in enumerate(desc_values): 
            if element !='Kājām': # ja sarkastā nav rakstīts, ka pārvietojas ar kājām, tad tiek pielikts elementam klāt numurs
                walk_transp.append(f'{element} {numbers_values.pop(0)}')
            else:
                walk_transp.append(element)
    else:
        for i, element in enumerate(desc_values): 
            if element !='Walk': # ja sarkastā nav rakstīts, ka pārvietojas ar kājām, tad tiek pielikts elementam klāt numurs
                walk_transp.append(f'{element} {numbers_values.pop(0)}')
            else:
                walk_transp.append(element)

    find = driver.find_element(By.XPATH, "/html/body/div[1]/div[3]/div[8]/div[9]/div/div/div[1]/div[2]/div/div[1]/div/div/div[1]/button") # atpakal poga
    find.click()
    time.sleep(5)

    find = driver.find_element(By.XPATH, "/html/body/div[1]/div[3]/div[8]/div[3]/div[1]/div[2]/div/div[2]/div/button") # aizver logu
    find.click()
    time.sleep(10)

    # savadam iegūto informāciju excel otrā darba lapā
    # tā kā sarakstiem ir atšķirīgi garumi, vispirms ievadam garāko sarakstu vērtības
    a=0
    datums = cycle_day
    iela_pietNos = h2_values[a]
    laiks_jabut = time_values[a]
    ws1_2['A'+str(start_row)].value=datums
    ws1_2['B'+str(start_row)].value=iela_pietNos
    ws1_2['C'+str(start_row)].value=laiks_jabut
    row_count_keeper+=1

    # tagad saraksti ir vienādā garumā un tiek ievadīti visi dati
    start_row=row_count_keeper # row_count_keeper uztur kurā rindā failā mēs atrodamies, tas ir atbildīgs informācijas atjaunošanu start_row elementam
    a=1
    for l in range(start_row, start_row + len(h2_values)-1):
        datums = cycle_day
        iela_pietNos = h2_values[a]
        laiks_jabut = time_values[a]
        iet_sabNos = walk_transp[a-1]
        sabTr_veids = el_style[a-1]
        ws1_2['A'+str(l)].value=datums
        ws1_2['B'+str(l)].value=iela_pietNos
        ws1_2['C'+str(l)].value=laiks_jabut
        ws1_2['D'+str(l)].value=iet_sabNos
        ws1_2['E'+str(l)].value=sabTr_veids
        a += 1
        row_count_keeper+=1
    start_row=row_count_keeper
    cycle += 2
driver.quit()

y-=1
dynamic_file_name = f"{x}_{y}kurss_{z}grupa.xlsx" # faila nosaukums mainās atkarībā no sākumā lietotāja ievadītās informācijas
wb1.save(dynamic_file_name)
wb1.close()